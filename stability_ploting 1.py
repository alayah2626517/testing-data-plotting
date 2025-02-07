#### Purpose: This script is for plotting testing data from excel, beware of the layout of data.
#### Author: Hsin-Yun Hung
#### Release time: 2025/02/

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.constants import *
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np
import os


def run_app():
    ## 建立窗口
    root = tk.Tk()
    root.title("Stability data plotting tool")
    root.geometry('380x380')
    root.iconbitmap('EG logo.ico')
    root.resizable(False, False)
   
    ## 輸入框: Batch
    label_frame = tk.LabelFrame(root, width=380, height=100, text="Step 1", bg="light blue", bd=10, relief='groove')
    label_frame.pack(padx=10, pady=5, fill="x")
    label_frame.pack_propagate(False)
    tk.Label(label_frame, text="Maximmum # of Batch ?", font=('Arial', 13), bg="light blue").grid(row=0, column=0, padx=60, pady=10)
    batch_num = tk.StringVar()
    batch_entry = tk.Spinbox(label_frame, from_=1, to=50, textvariable=batch_num, font=("Arial", 12))
    batch_entry.grid(row=1, column=0, padx=60, pady=15)

    ## 文件選擇按鈕
    def select_file():
        global file_path
        file_path = filedialog.askopenfilename(title="Select file", filetypes=[("All Files", "*.*")])
        if not file_path:
            return
        else:
            print(f"Successfully selecting: {file_path}")
    label_frame_2 = tk.LabelFrame(root, width=380, height=100, text="Step 2", bg="DarkOliveGreen2", bd=10, relief='groove')
    label_frame_2.pack(padx=10, pady=5, fill="x")
    tk.Label(label_frame_2, text="Beware of the datasets layout!", font=('Arial', 12), bg="DarkOliveGreen2").grid(row=0, column=0, padx=60, pady=10)
    file_path = None
    tk.Button(label_frame_2, text="Select file", command=select_file).grid(row=1, column=0, padx=60, pady=15)

    
     ## 文件導入按鈕
    def load_sheet():
        global file_path
        batch_value = batch_num.get()
        batch_value = int(batch_value)
        wb = load_workbook(file_path)
        sheet_name = wb.sheetnames
        folder_path = filedialog.askdirectory(title="選擇儲存路徑")
        for sheet in sheet_name:
            sheet = wb[sheet]
            data_total = []
            for data in sheet.iter_rows(min_row=1, max_row=batch_value+1, min_col=1, max_col=10, values_only=True):
                data_total.append(data)
            # print(data_total)
            
            # 定義各項變數
            condition = data_total[1][0]
            test_item = data_total[0][0]
            value_limit = None
            if data_total[0][2] == "±":
                low_limit = data_total[0][1]-data_total[0][3]
                upper_limit = data_total[0][1]+data_total[0][3]
                value_limit = np.linspace(low_limit, upper_limit, num=10)
            else:
                print("No specification defined.")            

            #製圖
            title = f"{condition}-{test_item}"
            datasets = data_total[2:]  # 數據本身
            plt.figure(figsize=(10, 6))
            for row in datasets:
                label = row[0]  # 每一行的標籤
                values = row[1:]  # 每一行的數據（跳過標籤）
                # x_axis = tuple(x for x in data_total[1][1:] if x is not None)
                x_axis = data_total[1][1:]
                plt.plot(x_axis, values, marker='o', linestyle='-', linewidth=2, alpha = 0.6, label=label)
            plt.title(title, fontsize=18, fontweight='bold')
            plt.xlabel("Time point (months)")
            plt.ylabel(test_item, fontsize=15)
            plt.yticks(value_limit, fontsize=12)
            plt.grid(True, linestyle='--', alpha=0.6)
            plt.legend()
            plt.grid(True)
            plt.savefig(f"{folder_path}/{title}.png", dpi=300)
        wb.close()
        if messagebox.askyesno("Exit", "All operations are completed. Do you want to exit?"):
        root.destroy()
    label_frame_3 = tk.LabelFrame(root, width=380, height=100, text="Step 2", bg="khaki1", bd=10, relief='groove')
    label_frame_3.pack(padx=10, pady=5, fill="x")
    tk.Button(label_frame_3, text="Load file", command=load_sheet).pack(pady=10)
    
    root.mainloop()
    


if __name__ == "__main__":
    run_app()
