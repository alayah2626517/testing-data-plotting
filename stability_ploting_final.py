#### Purpose: This script is for plotting testing data from excel, beware of the layout of data.
#### Author: Hsin-Yun Hung
#### Inition version: 2025/02/20

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.constants import *
import matplotlib.ticker as ticker
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np

def run_app():
    ## 建立窗口
    root = tk.Tk()
    root.title("Test results plotting tool")
    root.geometry('380x420')
    root.resizable(False, False)

    ## 輸入批次數量    
    label_frame = tk.LabelFrame(root, width=380, height=100, text="Step 1", bg="light blue", bd=10, relief='groove')
    label_frame.pack(padx=10, pady=5, fill="x")
    label_frame.pack_propagate(False)
    tk.Label(label_frame, text="Maximum # of Batch ?", font=('Arial', 13), bg="light blue").grid(row=0, column=0, padx=60, pady=10)
    batch_num = tk.StringVar()
    batch_entry = tk.Spinbox(label_frame, from_=1, to=100, textvariable=batch_num, font=("Arial", 12))
    batch_entry.grid(row=1, column=0, padx=70, pady=15)

    ## 文件選擇
    def select_file():
        global file_path
        file_path = filedialog.askopenfilename(title="Select file", filetypes=[("All Files", "*.*")])
        if not file_path:
            return
        else:
            print(f"Successfully selecting: {file_path}")
    label_frame_2 = tk.LabelFrame(root, width=380, height=100, text="Step 2", bg="DarkOliveGreen2", bd=10, relief='groove')
    label_frame_2.pack(padx=10, pady=5, fill="x")
    
    label_frame_2.grid_columnconfigure(0, weight=0)  # 第 1 列不會拉伸
    label_frame_2.grid_columnconfigure(1, weight=1)  # 第 2 列會根據容器大小拉伸
    label_frame_2.grid_columnconfigure(2, weight=0)  # 第 3 列不會拉伸    
    tk.Label(label_frame_2, text="Select the file: ", font=('Arial', 10), bg="DarkOliveGreen2").grid(row=0, column=1, padx=66, pady=5, sticky="ew")
    tk.Label(label_frame_2, text="Beware of the datasets layout!", font=('Arial', 11, 'bold'), bg="DarkOliveGreen2").grid(row=1, column=0, columnspan=3, padx=66, pady=5, sticky="ew")
    tk.Button(label_frame_2, text="Select file", command=select_file).grid(row=2, column=1, padx=55, pady=10, sticky="n")
    file_path = None
    
     ## 文件導入
    def load_sheet():
        global file_path
        batch_value = batch_num.get()
        batch_value = int(batch_value)
        wb = load_workbook(file_path)
        sheet_name = wb.sheetnames
        folder_path = filedialog.askdirectory(title="選擇儲存路徑")
        for sheet in sheet_name:
            sheet = wb[sheet]
            data_total = []
            for data in sheet.iter_rows(min_row=1, max_row=batch_value+2, min_col=1, max_col=10, values_only=True):
                data_total.append(data)

            ### 定義變數
            condition = data_total[1][0]
            test_item = data_total[0][0].split("(")[0]
            datasets = data_total[2:]  # 數據本身
            x_axis_raw = data_total[1][1:]  # time point
            x_axis = [int(x) if isinstance(x, (int, float)) else None for x in x_axis_raw]
            x_axis = [x for x in x_axis if x is not None]
            
            ## 找單位
            unit_ori = data_total[0][0].split("(")[1]
            end_idx = unit_ori.find(')')
            if end_idx != -1:
                chart_y_label = unit_ori[:end_idx]
                
            ## 訂出小數位數的function
            def get_decimal(value):
                if isinstance(value, (int, float)):
                    if '.' in str(value):
                        return len(str(value).split(".")[1])
                    return 0
                return 0
            max_decimal = 0
            for row in datasets:
                for value in row[1:]:
                    if value is not None:
                        max_decimal = max(max_decimal, get_decimal(value))
                
            ### 製作折線圖
            fig, ax = plt.subplots(1, 1, sharex='col', figsize=(10, 8))
            for row in datasets:
                label = row[0]  # 每一行的標籤
                values = row[1:]
                clean_x = []
                clean_y = []
                for x, y in zip(x_axis_raw, values):
                    if x is not None and isinstance(x, (int, float)):
                        clean_x.append(x)
                        clean_y.append(float(y) if isinstance(y, (int, float)) else np.nan)  # 每一行的數據（跳過標籤）
                #values = [float(v) if isinstance(v, (int, float)) else np.nan for v in row[1:]]  
                ax.plot(x_axis, values, marker='o', linestyle='-', linewidth=2, alpha = 0.6, label=label)

            ##定義y軸上下限
            value_limit, value_limit_none, lower_limit, upper_limit = None, None, None, None
            if data_total[0][2] == "±":
                lower_limit = data_total[0][1]-data_total[0][3]
                upper_limit = data_total[0][1]+data_total[0][3]
                num = 7
                value_limit = np.linspace(lower_limit*0.96, upper_limit*1.02, num=num)
                value_limit = [round(value, max_decimal) for value in value_limit]
                ax.set_yticks(value_limit)
                ax.axhline(y=lower_limit, color='#8B0000', linestyle='--', linewidth=1.5)
                ax.axhline(y=upper_limit, color='#8B0000', linestyle='--', linewidth=1.5)
                ax.text(x=-1.65, y=lower_limit, s=f"{lower_limit:.{max_decimal}f}", color='#8B0000', fontsize=12, ha='right', va='center')
                ax.text(x=-1.65, y=upper_limit, s=f"{upper_limit:.{max_decimal}f}", color='#8B0000', fontsize=12, ha='right', va='center')
            elif data_total[0][2] in ["<", "≦"]:
                upper_limit = data_total[0][3]
                distance = upper_limit*0.05
                num = 15
                value_limit = np.linspace(upper_limit-distance*(num-1), upper_limit*1.02, num=num)
                value_limit = [round(value, max_decimal) for value in value_limit]
                ax.set_yticks(value_limit)
                ax.axhline(y=upper_limit, color='#8B0000', linestyle='--', linewidth=1.5)
                ax.text(x=-0.8, y=upper_limit, s=f"{upper_limit:.{max_decimal}f}", color='#8B0000', fontsize=12, ha='right', va='center')
            elif data_total[0][2] in [">", "≧"]:
                lower_limit = data_total[0][3]
                distance = lower_limit*0.05
                num = 7
                value_limit = np.linspace(lower_limit*0.96, lower_limit+distance*(num-1), num=num)
                value_limit = [round(value, max_decimal) for value in value_limit]
                ax.set_yticks(value_limit)
                ax.axhline(y=lower_limit, color='#8B0000', linestyle='--', linewidth=1.5)
                ax.text(x=-1, y=lower_limit, s=f"{lower_limit:.{max_decimal}f}", color='#8B0000', fontsize=12, ha='right', va='center')
            else:
                all_values = [value for row in datasets for value in row[1:] if value is not None]
                min_value = min(all_values)
                max_value = max(all_values)                    
                value_limit_none = np.linspace(min_value*0.96, max_value*1.02)
                value_limit_none = [round(value, max_decimal) for value in value_limit_none]
                ax.set_yticks(value_limit_none)
            
            ### 製作表格
            data_rows = [row[1:] for row in datasets]  # 每一行的數據
            data_rows = [[f"{value:.{max_decimal}f}" if value is not None else None for value in row] for row in data_rows]
            bbox_height = 0.3 + (batch_value * 0.02)
            bbox_y = -0.5 - (batch_value * 0.01)
            data_labels = [row[0] for row in datasets]  # 提取批次label
            data_rows_labels = [[data_labels[i]] + data_rows[i] for i in range(len(data_rows))]
            table = ax.table(cellText=data_rows_labels, colLabels=['Lot'] + list(x_axis), loc='bottom', cellLoc='center', colLoc='center', bbox=[0, bbox_y, 1, bbox_height])
            table.auto_set_font_size(False)
            for (i, j), cell in table.get_celld().items():
                if i == 0:
                    cell.set_fontsize(11)
                elif j == 0:
                    cell.set_fontsize(6)
                else:
                    cell.set_fontsize(10)
            
            ### 整個表設計
            chart_title = f"{condition}-{test_item}"
            ax.set_title(chart_title, fontsize=18, fontweight='bold')
            ax.set_xlabel("Time point (months)")
            ax.set_xticks(x_axis)
            ax.yaxis.set_major_formatter(ticker.FormatStrFormatter(f"%.{max_decimal}f"))
            ax.set_ylabel(chart_y_label, fontsize=15)
            ax.grid(True, linestyle='--', alpha=0.6)
            ax.legend(loc='upper right', fontsize=10) ##, bbox_to_anchor=(-0.22, -0.45)
            ax.grid(True)
            plt.subplots_adjust(bottom=0.2)  # 調整底部的間距
            plt.tight_layout(pad=1.0)
            plt.savefig(f"{folder_path}/{chart_title}.png", dpi=300)
        
        wb.close()
        if messagebox.askyesno("Plotting complete", "All charts have been successfully created. Do you want to exit?"):
            root.quit()
    
    label_frame_3 = tk.LabelFrame(root, width=380, height=100, text="Step 3", bg="khaki1", bd=10, relief='groove')
    label_frame_3.pack(padx=10, pady=5, fill="x")
    label_frame_3.grid_columnconfigure(0, weight=1)
    label_frame_3.grid_columnconfigure(1, weight=1)
    label_frame_3.grid_columnconfigure(2, weight=1)
    tk.Button(label_frame_3, text="Load file", command=load_sheet).grid(row=0, column=1, padx=55, pady=10, sticky="n")
    tk.Label(label_frame_3, text="Select a folder after clicking the loading button.", font=('Arial', 11), bg="khaki1").grid(row=1, column=0, columnspan=3, padx=6, pady=10, sticky="ew")

    
    root.mainloop()
    


if __name__ == "__main__":
    run_app()
