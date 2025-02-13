#### Purpose: This script is for plotting testing data from excel, beware of the layout of data.
#### Author: Hsin-Yun Hung
#### Release time: 2025/02/

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.constants import *
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np
import os

# 擺好位置

def run_app():
    # 建立窗口
    root = tk.Tk()
    root.title("Data plotting tool")
    root.geometry('380x400')
    root.iconbitmap('EG logo.ico')
    root.resizable(False, False)

    # 輸入框: Batch
    label_frame = tk.LabelFrame(root, text="Step 1", bg="light blue", bd=10, relief='groove')
    label_frame.pack(padx=20, pady=5)
    label_frame.pack_propagate(False)
    tk.Label(label_frame, text="Maximmum # of Batch ?", font=('Arial', 14, 'bold'), bg="light blue").grid(row=0, column=0, padx=50, pady=10)
    batch_num = tk.StringVar()
    batch_entry = tk.Entry(label_frame, textvariable=batch_num)
    batch_entry.grid(row=1, column=0, padx=50, pady=20)

    # # 文件選擇按鈕
    label_frame_2 = tk.LabelFrame(root, width=380, height=100, text="Step 2", bg="DarkOliveGreen2", bd=10, relief='groove')
    label_frame_2.pack(padx=20, fill="x")
    file_path = None
    def select_file():
        global file_path
        file_path = filedialog.askopenfilename(title="Select file", filetypes=[("All Files", "*.*")])
        if not file_path:
            return
        else:
            print(f"Successfully selecting: {file_path}")
    tk.Button(label_frame_2, text="Select file", command=select_file).pack(pady=20)

    ## 文件導入按鈕
    label_frame_2 = tk.LabelFrame(root, width=380, height=100, text="Step 2", bg="DarkOliveGreen2", bd=10, relief='groove')
    label_frame_2.pack(padx=20, fill="x")
    def load_sheet():
        global file_path
        batch_value = batch_num.get()
        batch_value = int(batch_value)
        wb = load_workbook(file_path)
        sheet_name = wb.sheetnames
        for sheet in sheet_name:
            sheet = wb[sheet]
            data_total = []
            for data in sheet.iter_rows(min_row=1, max_row=batch_value+1, min_col=1, max_col=10, values_only=True):
                data_total.append(data)
            condition = data_total[1][0]
            test_item = data_total[0][0]
            temp_row = data_total[1]
            time_list = [value for value in temp_row if isinstance(value, (int, float))]
            T0 = time_list[0]
            Tlast = time_list[-1]
            #製圖

            # title = f"{condition}-{test_item}"

            print(T0, Tlast)
        wb.close()
    tk.Button(root, text="Load file", command=load_sheet).pack(pady=10)

    # wb = load_workbook(file)
    # sheet = wb[data_sheet]
    # return wb, sheet


    root.mainloop()

if __name__ == "__main__":
    run_app()
