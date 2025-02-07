#### Purpose: This script is for plotting testing data from excel, beware of the layout of data.
#### Author: Hsin-Yun Hung
#### Release time: 2025/02/

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.constants import *
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np
import os

# 擺好位置

def run_app():
    # 建立窗口
    root = tk.Tk()
    root.title("Data plotting tool")
    root.geometry('380x400')
    root.iconbitmap('EG logo.ico')
    root.resizable(False, False)

    # 輸入框: Batch
    label_frame = tk.LabelFrame(root, text="Step 1", width=380, heigh=100, bg="light blue", bd=10, relief='groove')
    label_frame.pack(padx=10, pady=10, fill="both", expand=True)
    tk.Label(root, text="Maximmum # of Batch ?", font=('Arial', 14, 'bold')).pack(pady=5)
    batch_num = tk.StringVar()
    tk.Entry(root, textvariable=batch_num).pack(pady=10)


    # # 文件選擇按鈕
    # label_frame = tk.LabelFrame(root, text="Step 2", width=380, heigh=100, bg="medium sea green", bd=10, relief='groove')
    # label_frame.pack(padx=200, pady=10)
    
    file_path = None
    def select_file():
        global file_path
        file_path = filedialog.askopenfilename(title="Select file", filetypes=[("All Files", "*.*")])
        if not file_path:
            return
        else:
            print(f"Successfully selecting: {file_path}")
    tk.Button(root, text="Select file", command=select_file).pack(pady=10)

    def load_sheet():
        global file_path
        batch_value = batch_num.get()
        batch_value = int(batch_value)
        wb = load_workbook(file_path)
        sheet_name = wb.sheetnames
        for sheet in sheet_name:
            sheet = wb[sheet]
            data_total = []
            for data in sheet.iter_rows(min_row=1, max_row=batch_value+1, min_col=1, max_col=10, values_only=True):
                data_total.append(data)
            condition = data_total[1][0]
            test_item = data_total[0][0]
            temp_row = data_total[1]
            time_list = [value for value in temp_row if isinstance(value, (int, float))]
            T0 = time_list[0]
            Tlast = time_list[-1]
            #製圖

            # title = f"{condition}-{test_item}"

            print(T0, Tlast)
        wb.close()
    tk.Button(root, text="Load file", command=load_sheet).pack(pady=10)

    # wb = load_workbook(file)
    # sheet = wb[data_sheet]
    # return wb, sheet


    root.mainloop()

if __name__ == "__main__":
    run_app()
